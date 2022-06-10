#!/bin/env/python3
import argparse
import os
import sys
import openpyxl
import paramiko

"""
get backend info for obsworkers by a obsworkers list file.
"""

#ArgumentParser
par = argparse.ArgumentParser()
par.add_argument("-f","--file",default = None,help = "obsworker list file.",required = True)
par.add_argument("-p","--passwd",default = None,help = "obsworker login passwd.",required = True)
args = par.parse_args()

curpath = os.getcwd()
workerlist = os.path.join(curpath,args.file)
print("work path:%s" %curpath)

sshuser = 'root'
sshpasswd = args.passwd
sshport = 22
sshcmd = r"cat /etc/sysconfig/obs-server | grep -i OBS_REPO_SERVER | grep -v '^#' | awk -F'=' '{print $2}' | sed 's/\.openeuler\.org:5252//g' | sed 's/\"//g'"

def sshclient(ip,port,user,passwd,cmd):
	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(ip,port,user,passwd,timeout = 5)
	stdin,stdout,stderr = ssh.exec_command(cmd)
	err_info = stderr.readlines()
	if len(err_info) > 0:
		print(err_info)
		return
	cmd_result = stdout.read()
	return cmd_result


def inputexcel(dict_msg):
	if os.path.exists('obsworker.xlsx'):
		os.remove('obsworker.xlsx')
	wb = openpyxl.Workbook()
	ws = wb.create_sheet('workers')
	ws.cell(1,1).value = 'IP'
	ws.cell(1,2).value = 'backend'
	n = 2
	for k,v in dict_msg.items():
		ws.cell(n,1).value = k
		ws.cell(n,2).value = v
		n = n+1
	wb.save('obsworker.xlsx')
	

def getbackendinfo(workerfile): 
	if not os.path.exists(workerfile):
		print("obsworker list file not exist,please check.")
		exit(1)
	backend_dict = {}
	with open(workerfile,'r',encoding = 'utf-8') as f:
		for IP in f.readlines():
			IP = IP.replace('\n','')
			print("start check: %s" %IP)
			backend = sshclient(IP,22,sshuser,sshpasswd,sshcmd)
			if not backend:
				print("check failed: %s" %IP)
			print("check result: ",backend)
			backend_dict[IP] = backend
	inputexcel(backend_dict)
	

getbackendinfo(workerlist)

